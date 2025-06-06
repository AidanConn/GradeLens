from fastapi import FastAPI, APIRouter, File, UploadFile, HTTPException, Request, Response, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from typing import List
import os
import uuid
import json
from pathlib import Path
from datetime import datetime
from filelock import FileLock
import csv
from io import StringIO
import re
import math

# Initialize FastAPI app
app = FastAPI(title="GradeLens API")

# Configure CORS
origins = [
    "http://localhost:3000",  # React app
    "http://localhost:5173",  # Dev - With out docker
    "http://localhost:8000",  # FastAPI app
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    EXCEL_EXPORT_AVAILABLE = False

@app.get("/")
async def root(request: Request, response: Response):
    session_id = get_session_id(request, response)
    return {"message": "Welcome to your FastAPI application!", "session_id": session_id}

router = APIRouter()

UPLOAD_DIR = Path("/app/uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

def get_session_id(request: Request, response: Response):
    session_id = request.cookies.get("session_id")
    if not session_id:
        session_id = request.headers.get("X-Session-ID")
    if not session_id:
        session_id = str(uuid.uuid4())
        response.set_cookie(
            key="session_id",
            value=session_id,
            httponly=True,
            path="/",
            samesite="lax",
            max_age=604800  
        )
    session_files_dir = UPLOAD_DIR / session_id / "files"
    session_runs_dir = UPLOAD_DIR / session_id / "runs"
    session_files_dir.mkdir(parents=True, exist_ok=True)
    session_runs_dir.mkdir(parents=True, exist_ok=True)
    return session_id

def parse_csv_lines(lines: List[str]) -> List[List[str]]:
    """
    Convert a list of strings (lines) into a list of CSV rows.
    """
    csv_text = "\n".join(lines)
    reader = csv.reader(StringIO(csv_text))
    return list(reader)

def parse_sec_file(lines: List[str]) -> dict:
    """
    Parses a .sec file.
    Expected formats:
      - CSV Format: Header row with comma separation:
           course_code, semester, [credit_hours]
      - Plain Text Format: A single header field containing whitespace-separated tokens,
           e.g. "COMSC110.01F22  4.0" (course code then credit hours)
      Subsequent rows (CSV format) must contain: name, student_id, grade

    Returns a dictionary with the course info nested under "course" and the list of student records.
    """
    rows = parse_csv_lines(lines)
    if not rows:
        raise ValueError("Empty file")

    header = rows[0]
    if len(header) == 1:
        # Plain text header; split on whitespace.
        parts = header[0].split()
        if not parts:
            raise ValueError(f"Header line is empty after splitting: {header[0]}")
        course_name = parts[0]
        credit_hours = None
        if len(parts) > 1:
            try:
                credit_hours = float(parts[1])
            except ValueError as e:
                raise ValueError(f"Unable to convert credit value '{parts[1]}' to float: {e}")
        # No semester is available in this header.
        course_info = {"name": course_name, "credit_hours": credit_hours}
    else:
        # CSV header format with multiple columns.
        course_name = header[0].strip() if len(header) > 0 else "Unknown Course"
        semester = header[1].strip() if len(header) > 1 else None
        credit_hours = None
        if len(header) > 2:
            try:
                credit_hours = float(header[2].strip())
            except ValueError as e:
                raise ValueError(f"Unable to convert credit value '{header[2]}' to float: {e}")
        course_info = {"name": course_name, "credit_hours": credit_hours}
        if semester:
            course_info["semester"] = semester

    students = []
    for i, row in enumerate(rows[1:], start=2):
        if len(row) < 3:
            raise ValueError(f"Row {i} does not have enough columns: {row}")
        name = row[0].strip()
        student_id = row[1].strip()
        grade = row[2].strip()
        students.append({"name": name, "student_id": student_id, "grade": grade})

    return {"course": course_info, "students": students}


def parse_grp_file(lines: List[str]) -> dict:
    """
    Parses a .grp file.
    Expected format (plain text):
      - First non-empty line: group name (e.g., COMSC400)
      - Subsequent non-empty lines: section file names (e.g., COMSC401.01F22.sec, etc.)
    Returns a dictionary with the group name and section filenames.
    """
    clean_lines = [line.strip() for line in lines if line.strip()]
    if not clean_lines:
        raise ValueError("Empty file")
    
    group_name = clean_lines[0]

    # Remove the .sec from the section names
    sections = []
    for i in range(1, len(clean_lines)):
        sections.append(clean_lines[i].replace(".sec", ""))

    return {"group_name": group_name, "sections": sections}

@router.post("/upload_sec_grp/")
async def upload_sec_grp_files(
    files: List[UploadFile] = File(...),
    session_id: str = Depends(get_session_id)
):
    """
    Uploads and processes multiple .SEC or .GRP files.
    - For .sec files, expects a header as defined in parse_sec_file.
    - For .grp files, expects plain text with the first non-empty line as the course and the rest as section filenames.
    The parsed data is stored as JSON.
    """
    allowed_extensions = {".sec", ".grp"}
    session_files_dir = UPLOAD_DIR / session_id / "files"
    results = []

    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid file type '{ext}' for file '{file.filename}'. Only .SEC and .GRP files are allowed."
            )

        file_path = session_files_dir / file.filename
        content = await file.read()
        try:
            lines = content.decode("utf-8").splitlines()
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail=f"Unable to decode file '{file.filename}'.")

        if not lines:
            raise HTTPException(status_code=400, detail=f"Uploaded file '{file.filename}' is empty.")

        try:
            if ext == ".sec":
                parsed_data = parse_sec_file(lines)
            elif ext == ".grp":
                parsed_data = parse_grp_file(lines)
            else:
                raise HTTPException(status_code=400, detail=f"Unsupported file type '{ext}'.")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing file '{file.filename}': {str(e)}")

        json_path = file_path.with_suffix(".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(parsed_data, f, indent=2)

        results.append({
            "message": "File processed successfully",
            "original_filename": file.filename,
            "stored_as": json_path.name,
            "parsed_data": parsed_data
        })

    return {"files": results}

@router.post("/upload_run/")
async def upload_run_file(
    run_file: UploadFile = File(...),
    session_id: str = Depends(get_session_id)
):
    """
    Uploads and parses a .RUN file without performing calculations.
    The run file is stored and associated files are identified.
    """
    ext = os.path.splitext(run_file.filename)[1].lower()
    if ext != ".run":
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type '{ext}' for run file '{run_file.filename}'."
        )
    run_id = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    run_dir = UPLOAD_DIR / session_id / "runs" / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    run_file_path = run_dir / run_file.filename
    content = await run_file.read()
    with open(run_file_path, "wb") as f:
        f.write(content)

    # Parse the run file to identify associated files
    associated_files = parse_run_file_for_associated_files(run_file_path)
    with open(run_dir / "associated_files.json", "w") as f:
        json.dump(associated_files, f, indent=2)
    
    return {
        "message": "Run file uploaded and parsed successfully",
        "run_id": run_id,
        "run_name": associated_files["run_name"],
        "associated_files": associated_files["grp_files"]
    }

@router.post("/runs/{run_id}/calculate")
async def calculate_run(
    run_id: str, 
    session_id: str = Depends(get_session_id)
):
    """
    Performs calculations on a previously uploaded run file.
    """
    run_dir = UPLOAD_DIR / session_id / "runs" / run_id
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")
    
    # Find the run file
    run_files = list(run_dir.glob("*.run"))
    if not run_files:
        raise HTTPException(status_code=404, detail=f"No run file found in run {run_id}.")
    
    run_file_path = run_files[0]
    
    # Load associated files
    associated_files_path = run_dir / "associated_files.json"
    if not associated_files_path.exists():
        raise HTTPException(status_code=404, detail=f"Associated files data not found for run {run_id}.")
    
    with open(associated_files_path, "r") as f:
        associated_files = json.load(f)
    
    # Process the run file and generate calculations
    calc_results = process_run_file(run_file_path, associated_files, session_id)
    
    # Save calculations
    calc_file_path = run_dir / "calculations.json"
    lock = FileLock(str(calc_file_path) + ".lock")
    with lock:
        with open(calc_file_path, "w") as f:
            json.dump(calc_results, f, indent=2)
    
    return {
        "message": "Run calculations completed",
        "run_id": run_id,
        "calculation_results": calc_results
    }


@router.get("/runs/")
async def list_runs(session_id: str = Depends(get_session_id)):
    """
    Lists all available runs for the current session.
    """
    session_runs_dir = UPLOAD_DIR / session_id / "runs"
    if not session_runs_dir.exists():
        return {"runs": []}
    
    runs = []
    for run_dir in session_runs_dir.iterdir():
        if not run_dir.is_dir():
            continue
        
        run_id = run_dir.name
        run_info = {"run_id": run_id}
        
        # Get run name and other metadata if available
        associated_files_path = run_dir / "associated_files.json"
        if associated_files_path.exists():
            with open(associated_files_path, "r") as f:
                associated_files = json.load(f)
                run_info["run_name"] = associated_files.get("run_name", "Unnamed Run")
                run_info["associated_files"] = associated_files.get("grp_files", [])
        
        # Check if calculations exist
        calc_file_path = run_dir / "calculations.json"
        run_info["calculations_exist"] = calc_file_path.exists()
        
        # Get creation time based on directory creation time
        creation_time = datetime.fromtimestamp(run_dir.stat().st_ctime)
        run_info["created_at"] = creation_time.isoformat()
        
        runs.append(run_info)
    
    # Sort runs by creation time (newest first)
    runs.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    
    return {"runs": runs}

@router.get("/runs/{run_id}/calculations")
async def get_run_calculations(
    run_id: str, 
    session_id: str = Depends(get_session_id)
):
    """
    Retrieves the calculation results for a specific run.
    """
    run_dir = UPLOAD_DIR / session_id / "runs" / run_id
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")
    
    calc_file_path = run_dir / "calculations.json"
    if not calc_file_path.exists():
        raise HTTPException(status_code=404, detail=f"Calculations not found for run {run_id}.")
    
    with open(calc_file_path, "r", encoding="utf-8") as f:
        calculations = json.load(f)
    
    return {
        "run_id": run_id,
        "calculations": calculations
    }

@router.get("/runs/{run_id}/status")
async def get_run_status(
    run_id: str, 
    session_id: str = Depends(get_session_id)
):
    """
    Checks if calculations have been performed for a specific run.
    """
    run_dir = UPLOAD_DIR / session_id / "runs" / run_id
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")
    
    calc_file_path = run_dir / "calculations.json"
    calculations_exist = calc_file_path.exists()
    
    # Get associated files info
    associated_files_path = run_dir / "associated_files.json"
    associated_files = None
    if associated_files_path.exists():
        with open(associated_files_path, "r") as f:
            associated_files = json.load(f)
    
    return {
        "run_id": run_id,
        "calculations_exist": calculations_exist,
        "run_info": associated_files
    }

@router.get("/runs/comparison")
async def compare_runs(session_id: str = Depends(get_session_id)):
    session_runs_dir = UPLOAD_DIR / session_id / "runs"
    work_map, good_map = {}, {}
    if session_runs_dir.exists():
        for run_dir in session_runs_dir.iterdir():
            if not run_dir.is_dir(): continue
            run_id = run_dir.name
            calc_file = run_dir / "calculations.json"
            if not calc_file.exists(): continue
            with open(calc_file, "r", encoding="utf-8") as f:
                calc = json.load(f)
            for s in calc.get("improvement_lists", {}).get("work_list", []):
                name = s.get("name")
                if name:
                    e = work_map.setdefault(name, {"count":0,"runs":[]})
                    e["count"] += 1
                    e["runs"].append(run_id)
            for s in calc.get("improvement_lists", {}).get("good_list", []):
                name = s.get("name")
                if name:
                    e = good_map.setdefault(name, {"count":0,"runs":[]})
                    e["count"] += 1
                    e["runs"].append(run_id)
    return {"work_list_comparison": work_map, "good_list_comparison": good_map}

# Helper functions for parsing and processing files
def parse_run_file_for_associated_files(run_file_path: Path) -> dict:
    """
    Parses a .run file to extract the run name and associated .GRP files.
    
    Expected format:
    - First line: Run name (e.g., "FIRSTRUN")
    - Subsequent lines: GRP file names (e.g., "COMSC100.GRP")
    
    Returns a dictionary with the run name and associated files.
    """
    try:
        with open(run_file_path, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]
            
        if not lines:
            raise ValueError("Empty run file")
            
        run_name = lines[0]
        grp_files = []
        
        for i in range(1, len(lines)):
            file_name = lines[i]
            # Ensure .GRP extension if not already present
            if not file_name.lower().endswith('.grp'):
                file_name = f"{file_name}.GRP"
            grp_files.append(file_name)
            
        return {
            "run_name": run_name,
            "grp_files": grp_files
        }
    except Exception as e:
        raise ValueError(f"Error parsing run file: {str(e)}")

def process_run_file(run_file_path: Path, associated_files: dict, session_id: str) -> dict:
    """
    Process a run file using the associated GRP files.
    
    Args:
        run_file_path: Path to the .run file
        associated_files: Dictionary containing run_name and grp_files
        session_id: User session ID for finding related files
        
    Returns:
        A dictionary with calculation results
    """
    run_name = associated_files.get("run_name", "Unnamed Run")
    grp_files = associated_files.get("grp_files", [])
    
    # Initialize results
    results = {
        "run_name": associated_files.get("run_name", "Unnamed Run"),
        "groups": [],
        "courses": {},  # Will store data grouped by actual course codes
        "class_types": {},
        "students": {},
        "improvement_lists": {
            "work_list": [],  # Students needing improvement
            "good_list": []   # Students performing well (B+ or better)
        }
    }
    
    session_files_dir = UPLOAD_DIR / session_id / "files"
    
    # RWU GPA scale values
    grade_values = {
        "A": 4.0, "A-": 3.67,
        "B+": 3.33, "B": 3.0, "B-": 2.67,
        "C+": 2.33, "C": 2.0, "C-": 1.67,
        "D+": 1.33, "D": 1.0, "D-": 0.67,
        "F": 0.0
    }
    
    # Define grade categories for consistent distribution
    grade_categories = {
        "A": ["A", "A-"],
        "B": ["B+", "B", "B-"],
        "C": ["C+", "C", "C-"],
        "D": ["D+", "D", "D-"],
        "F": ["F"],
        "W": ["W"],
        "Other": []  # Will capture any grades not listed above
    }
    
    # Helper function to determine course type based on course code
    def get_course_type(course_code):
        # Extract course level from course code (e.g. COMSC100 -> 100-level)
        match = re.search(r'(\d{3})', course_code)
        if match:
            level = match.group(1)[0]  # First digit of the course number
            if level == '1':
                return "100-level"
            elif level == '2':
                return "200-level"
            elif level == '3':
                return "300-level"
            elif level == '4':
                return "400-level"
        return "other"
    
    # Helper function to extract course code from section name
    def extract_course_code(section_name):
        # Match patterns like COMSC401 from COMSC401.01F18
        match = re.match(r'^([A-Z]+\d{3})', section_name)
        if match:
            return match.group(1)
        return section_name
    
    # Process each GRP file
    for grp_file in grp_files:
        grp_path = session_files_dir / grp_file
        grp_json_path = grp_path.with_suffix(".json")
        
        if not grp_json_path.exists():
            results["groups"].append({
                "group_file": grp_file,
                "error": f"Could not find processed JSON for {grp_file}"
            })
            continue
            
        # Load the GRP data
        with open(grp_json_path, "r", encoding="utf-8") as f:
            grp_data = json.load(f)
            
        group_name = grp_data.get("group_name", "Unknown Group")
        sections = grp_data.get("sections", [])
        
        # Deduplicate sections for this group
        unique_sections = list(dict.fromkeys(sections))
        
        group_results = {
            "group_name": group_name,
            "sections": unique_sections,  # Only unique sections
            "courses": []
        }
        
        # Group sections by actual course code
        sections_by_course = {}
        for section in unique_sections:
            course_code = extract_course_code(section)
            if course_code not in sections_by_course:
                sections_by_course[course_code] = []
            sections_by_course[course_code].append(section)
            
            # Initialize course in results if not already present
            if course_code not in results["courses"]:
                course_type = get_course_type(course_code)
                results["courses"][course_code] = {
                    "course_name": course_code,
                    "course_type": course_type,
                    "total_students": 0,
                    "sections": [],
                    "grade_distribution": {category: 0 for category in grade_categories},
                    "detailed_grade_distribution": {grade: 0 for grade in grade_values.keys()},
                    "average_gpa": 0.0,
                    "student_performance": []
                }
                
                # Initialize class type metrics if not already present
                if course_type not in results["class_types"]:
                    results["class_types"][course_type] = {
                        "total_students": 0,
                        "grade_distribution": {category: 0 for category in grade_categories},
                        "detailed_grade_distribution": {grade: 0 for grade in grade_values.keys()},
                        "average_gpa": 0.0,
                        "courses": []
                    }
                if course_code not in results["class_types"][course_type]["courses"]:
                    results["class_types"][course_type]["courses"].append(course_code)
        
        group_results["courses"] = list(sections_by_course.keys())
        
        # --- Group-internal z-score for sections ---
        # Collect section GPAs for this group (unique only)
        section_gpa_map = {}
        for course_code, course_sections in sections_by_course.items():
            for section in course_sections:
                section_file = f"{section}.json"
                section_path = session_files_dir / section_file
                
                if not section_path.exists():
                    continue
                    
                # Load the section data
                with open(section_path, "r", encoding="utf-8") as f:
                    section_data = json.load(f)
                    
                students = section_data.get("students", [])
                section_info = section_data.get("course", {})
                
                # Get credit hours for this section
                credit_hours = section_info.get("credit_hours", 3.0)  # Default to 3.0 if not specified
                if credit_hours is None:
                    credit_hours = 3.0  # Default if explicitly None
                    
                # Calculate section GPA (weighted by credit hours)
                section_total_points = 0.0
                section_graded_students = 0
                
                for student in students:
                    grade = student.get("grade", "").upper()
                    if grade in grade_values:
                        grade_point = grade_values[grade]
                        section_total_points += grade_point * credit_hours
                        section_graded_students += 1
                
                if section_graded_students > 0 and credit_hours > 0:
                    avg_gpa = round(section_total_points / (section_graded_students * credit_hours), 2)
                else:
                    avg_gpa = 0.0
                
                section_gpa_map[section] = avg_gpa
        
        # Compute mean and stddev for group-internal z-score
        gpa_values = list(section_gpa_map.values())
        if gpa_values:
            mean_gpa = sum(gpa_values) / len(gpa_values)
            stddev_gpa = math.sqrt(sum((g - mean_gpa) ** 2 for g in gpa_values) / len(gpa_values)) if len(gpa_values) > 1 else 0.0
        else:
            mean_gpa = stddev_gpa = 0.0
        
        # Assign group_z_score to each section in this group
        group_section_z_scores = {}
        for section, gpa in section_gpa_map.items():
            if stddev_gpa > 0:
                group_section_z_scores[section] = round((gpa - mean_gpa) / stddev_gpa, 2)
            else:
                group_section_z_scores[section] = 0.0
        
        group_results["section_gpas"] = section_gpa_map
        group_results["section_group_z_scores"] = group_section_z_scores
        group_results["group_section_z_mean"] = round(mean_gpa, 2)
        group_results["group_section_z_stddev"] = round(stddev_gpa, 2)
        
        results["groups"].append(group_results)
        
        # Process each section and add to appropriate course
        for course_code, course_sections in sections_by_course.items():
            course_type = get_course_type(course_code)
            course_data = results["courses"][course_code]
            
            # Track course-level metrics
            total_grade_points = 0.0
            total_credit_hours = 0.0
            total_graded_students = 0
            
            section_gpas = []

            for section in course_sections:
                section_file = f"{section}.json"
                section_path = session_files_dir / section_file
                
                if not section_path.exists():
                    course_data["sections"].append({
                        "section_name": section,
                        "error": f"Could not find processed JSON for section {section}"
                    })
                    continue
                    
                # Load the section data
                with open(section_path, "r", encoding="utf-8") as f:
                    section_data = json.load(f)
                    
                students = section_data.get("students", [])
                section_info = section_data.get("course", {})
                
                # Get credit hours for this section
                credit_hours = section_info.get("credit_hours", 3.0)  # Default to 3.0 if not specified
                if credit_hours is None:
                    credit_hours = 3.0  # Default if explicitly None
                    
                section_results = {
                    "section_name": section,
                    "credit_hours": credit_hours,
                    "student_count": len(students),
                    "grade_distribution": {category: 0 for category in grade_categories},
                    "detailed_grade_distribution": {grade: 0 for grade in grade_values.keys()},
                    "average_gpa": 0.0,
                    "students": []  # Track individual student performance in this section
                }
                
                section_total_points = 0.0
                section_graded_students = 0
                
                # Process student grades
                for student in students:
                    student_name = student.get("name", "Unknown")
                    student_id = student.get("student_id", "000000")
                    grade = student.get("grade", "").upper()
                    
                    # Create student record for this section
                    student_record = {
                        "name": student_name,
                        "id": student_id,
                        "grade": grade,
                        "grade_point": grade_values.get(grade, None),
                        "section": section,
                        "course": course_code,
                        "course_type": course_type,
                        "credit_hours": credit_hours
                    }
                    
                    # Add to section students list
                    section_results["students"].append(student_record)
                    
                    # Add to course student performance list
                    course_data["student_performance"].append(student_record)
                    
                    # Update cross-student tracking
                    if student_id not in results["students"]:
                        results["students"][student_id] = {
                            "name": student_name,
                            "id": student_id,
                            "courses": [],
                            "total_grade_points": 0.0,
                            "total_credit_hours": 0.0,
                            "total_courses": 0,
                            "gpa": 0.0
                        }
                    
                    # Update grade distributions
                    if grade in grade_values:
                        # Track in detailed distribution for section and course
                        section_results["detailed_grade_distribution"][grade] += 1
                        course_data["detailed_grade_distribution"][grade] += 1
                        results["class_types"][course_type]["detailed_grade_distribution"][grade] += 1
                        
                        # Add to category totals (A, B, C, D, F)
                        for category, grades in grade_categories.items():
                            if grade in grades:
                                section_results["grade_distribution"][category] += 1
                                course_data["grade_distribution"][category] += 1
                                results["class_types"][course_type]["grade_distribution"][category] += 1
                                break
                        
                        # Calculate GPA with credit hours weighting
                        grade_point = grade_values[grade]
                        weighted_grade_point = grade_point * credit_hours
                        
                        section_total_points += weighted_grade_point
                        section_graded_students += 1
                        
                        total_grade_points += weighted_grade_point
                        total_credit_hours += credit_hours
                        total_graded_students += 1
                        
                        # Update student cross-course records
                        results["students"][student_id]["courses"].append({
                            "course": course_code,
                            "section": section,
                            "grade": grade,
                            "grade_point": grade_point,
                            "credit_hours": credit_hours,
                            "weighted_points": weighted_grade_point,
                            "course_type": course_type
                        })
                        results["students"][student_id]["total_grade_points"] += weighted_grade_point
                        results["students"][student_id]["total_credit_hours"] += credit_hours
                        results["students"][student_id]["total_courses"] += 1
                    
                    elif grade == "W":
                        section_results["grade_distribution"]["W"] += 1
                        course_data["grade_distribution"]["W"] += 1
                        results["class_types"][course_type]["grade_distribution"]["W"] += 1
                    
                    else:
                        # Handle any other grade (NP, I, etc.)
                        section_results["grade_distribution"]["Other"] += 1
                        course_data["grade_distribution"]["Other"] += 1
                        results["class_types"][course_type]["grade_distribution"]["Other"] += 1
                
                # Calculate section GPA - use credit hours if available
                if section_graded_students > 0 and credit_hours > 0:
                    section_results["average_gpa"] = round(section_total_points / (section_graded_students * credit_hours), 2)
                else:
                    section_results["average_gpa"] = 0.0
                
                section_gpas.append(section_results["average_gpa"])
                course_data["total_students"] += section_results["student_count"]
                results["class_types"][course_type]["total_students"] += section_results["student_count"]
                course_data["sections"].append(section_results)
            
            # Calculate course GPA using credit hours weighting
            if total_graded_students > 0 and total_credit_hours > 0:
                course_data["average_gpa"] = round(total_grade_points / total_credit_hours, 2)

            # --- Z-SCORE CALCULATION FOR SECTIONS ---
            # Compute mean and stddev for section GPAs
            if section_gpas:
                mean_gpa = sum(section_gpas) / len(section_gpas)
                stddev_gpa = math.sqrt(sum((g - mean_gpa) ** 2 for g in section_gpas) / len(section_gpas)) if len(section_gpas) > 1 else 0.0
                # Assign z-score to each section
                for section in course_data["sections"]:
                    if stddev_gpa > 0:
                        section["z_score"] = round((section["average_gpa"] - mean_gpa) / stddev_gpa, 2)
                    else:
                        section["z_score"] = 0.0
            # --- END Z-SCORE CALCULATION ---

    # --- GROUP Z‑SCORE CALCULATION FOR GROUPS ---
    group_gpas = []
    for grp in results["groups"]:
        if "courses" not in grp:
            continue
        pts = 0.0
        studs = 0
        course_gpas = []
        course_gpa_map = {}
        for code in grp["courses"]:
            c = results["courses"].get(code)
            if c:
                pts += c["average_gpa"] * c["total_students"]
                studs += c["total_students"]
                course_gpas.append(c["average_gpa"])
                course_gpa_map[code] = c["average_gpa"]
        grp_gpa = round(pts / studs, 2) if studs > 0 else 0.0
        grp["average_gpa"] = grp_gpa
        group_gpas.append(grp_gpa)

        # --- Z-SCORE OF COURSES WITHIN THIS GROUP ---
        # Only for the courses in this group
        course_gpas = [results["courses"][code]["average_gpa"] for code in grp["courses"] if code in results["courses"]]
        if course_gpas:
            mean = sum(course_gpas) / len(course_gpas)
            std = math.sqrt(sum((g - mean) ** 2 for g in course_gpas) / len(course_gpas)) if len(course_gpas) > 1 else 0.0
            grp["course_z_scores"] = {}
            for code in grp["courses"]:
                c = results["courses"].get(code)
                if c:
                    grp["course_z_scores"][code] = round((c["average_gpa"] - mean) / std, 2) if std > 0 else 0.0

    # Calculate overall mean and std‑dev of group GPAs
    if group_gpas:
        mean_gpa = sum(group_gpas) / len(group_gpas)
        std_dev = math.sqrt(
            sum((g - mean_gpa) ** 2 for g in group_gpas) / len(group_gpas)
        ) if len(group_gpas) > 1 else 0.0
    else:
        mean_gpa = std_dev = 0.0

    # Assign Z‑score to each group
    for grp, g in zip(results["groups"], group_gpas):
        grp["z_score"] = round((g - mean_gpa) / std_dev, 2) if std_dev > 0 else 0.0

    # Store group comparison summary
    results["group_comparison"] = {
        "mean": round(mean_gpa, 2),
        "std_dev": round(std_dev, 2)
    }
    # --- END GROUP Z‑SCORE CALCULATION ---

    # --- COURSE G-SCORE CALCULATION WITHIN EACH LEVEL ---
    for course_type, type_data in results["class_types"].items():
        # Gather all courses of this type
        course_codes = type_data.get("courses", [])
        course_gpas = []
        for code in course_codes:
            course = results["courses"].get(code)
            if course:
                course_gpas.append(course["average_gpa"])
        # Calculate mean and stddev
        if course_gpas:
            mean_gpa = sum(course_gpas) / len(course_gpas)
            stddev_gpa = math.sqrt(sum((g - mean_gpa) ** 2 for g in course_gpas) / len(course_gpas)) if len(course_gpas) > 1 else 0.0
        else:
            mean_gpa = stddev_gpa = 0.0
        # Assign G-score to each course in this level
        for code in course_codes:
            course = results["courses"].get(code)
            if course:
                if stddev_gpa > 0:
                    course["g_score"] = round((course["average_gpa"] - mean_gpa) / stddev_gpa, 2)
                else:
                    course["g_score"] = 0.0
        # Optionally, store mean/stddev for reference
        type_data["g_score_mean"] = round(mean_gpa, 2)
        type_data["g_score_stddev"] = round(stddev_gpa, 2)
    # --- END COURSE G-SCORE CALCULATION ---

    # Calculate class type GPAs using credit hours weighting
    for course_type, type_data in results["class_types"].items():
        total_points = 0.0
        total_credits = 0.0
        
        # Recalculate from course data for accurate weighting
        for course_code, course_data in results["courses"].items():
            if course_data["course_type"] == course_type:
                # Sum up all grade points and credit hours for this course type
                for student_record in course_data["student_performance"]:
                    if student_record.get("grade") in grade_values:
                        total_points += student_record["grade_point"] * student_record["credit_hours"]
                        total_credits += student_record["credit_hours"]
        
        if total_credits > 0:
            type_data["average_gpa"] = round(total_points / total_credits, 2)
    
    # Calculate per-student GPA using credit hours weighting
    for student_id, student_data in results["students"].items():
        if student_data["total_credit_hours"] > 0:
            student_data["gpa"] = round(student_data["total_grade_points"] / student_data["total_credit_hours"], 2)
            
            # Categorize student performance
            if student_data["gpa"] < 2.0:  # Below C average
                results["improvement_lists"]["work_list"].append({
                    "id": student_id,
                    "name": student_data["name"],
                    "gpa": student_data["gpa"],
                    "courses": student_data["courses"]
                })
            elif student_data["gpa"] >= 3.3:  # B+ or better
                results["improvement_lists"]["good_list"].append({
                    "id": student_id,
                    "name": student_data["name"],
                    "gpa": student_data["gpa"],
                    "courses": student_data["courses"]
                })
    
    # Sort performance lists by GPA
    results["improvement_lists"]["work_list"].sort(key=lambda x: x["gpa"])
    results["improvement_lists"]["good_list"].sort(key=lambda x: x["gpa"], reverse=True)
    
    # --- GROUP Z‑SCORE CALCULATION FOR GROUPS ---
    # Compute each group's aggregate GPA (weighted by total students)
    group_gpas = []
    for grp in results["groups"]:
        if "courses" not in grp:
            continue
        pts = 0.0
        studs = 0
        for code in grp["courses"]:
            c = results["courses"].get(code)
            if c:
                pts += c["average_gpa"] * c["total_students"]
                studs += c["total_students"]
        grp_gpa = round(pts / studs, 2) if studs > 0 else 0.0
        grp["average_gpa"] = grp_gpa
        group_gpas.append(grp_gpa)

    # Calculate overall mean and std‑dev of group GPAs
    if group_gpas:
        mean_gpa = sum(group_gpas) / len(group_gpas)
        std_dev = math.sqrt(
            sum((g - mean_gpa) ** 2 for g in group_gpas) / len(group_gpas)
        ) if len(group_gpas) > 1 else 0.0
    else:
        mean_gpa = std_dev = 0.0

    # Assign Z‑score to each group
    for grp, g in zip(results["groups"], group_gpas):
        grp["z_score"] = round((g - mean_gpa) / std_dev, 2) if std_dev > 0 else 0.0

    # Store group comparison summary
    results["group_comparison"] = {
        "mean": round(mean_gpa, 2),
        "std_dev": round(std_dev, 2)
    }
    # --- END GROUP Z‑SCORE CALCULATION ---

    # Calculate overall statistics
    unique_student_ids = set(results["students"].keys())
    # Count unique students across all courses
    total_students = len(unique_student_ids)
    
    overall_grades = {category: 0 for category in grade_categories}
    detailed_grades = {grade: 0 for grade in grade_values.keys()}
    
    # Convert courses dictionary to a list for the results
    results["course_list"] = []
    for course_code, course_data in results["courses"].items():
        results["course_list"].append(course_data)
        
        # Accumulate grade distributions for summary
        for grade, count in course_data["detailed_grade_distribution"].items():
            if grade in detailed_grades:
                detailed_grades[grade] += count
        
        for category, count in course_data["grade_distribution"].items():
            if category in overall_grades:
                overall_grades[category] += count
    
    # Calculate overall GPA with credit hour weighting
    overall_total_points = 0.0
    overall_total_credits = 0.0
    
    # Use the student-level data to calculate the overall GPA
    for student_id, student_data in results["students"].items():
        overall_total_points += student_data["total_grade_points"]
        overall_total_credits += student_data["total_credit_hours"]
    
    overall_gpa = 0.0
    if overall_total_credits > 0:
        overall_gpa = round(overall_total_points / overall_total_credits, 2)
    
    results["summary"] = {
        "total_students": total_students,
        "grade_distribution": overall_grades,
        "detailed_grade_distribution": detailed_grades,
        "overall_gpa": overall_gpa,
        "total_credit_hours": overall_total_credits
    }
    
    return results

# Export to Excel functionality
@router.get("/runs/{run_id}/export")
async def export_run_to_excel(
    run_id: str,
    session_id: str = Depends(get_session_id)
):
    """
    Exports calculation results for a specific run to Excel format.
    """
    if not EXCEL_EXPORT_AVAILABLE:
        raise HTTPException(
            status_code=501, 
            detail="Excel export functionality is not available. Please install openpyxl package."
        )
    
    run_dir = UPLOAD_DIR / session_id / "runs" / run_id
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail=f"Run {run_id} not found.")
    
    calc_file_path = run_dir / "calculations.json"
    if not calc_file_path.exists():
        raise HTTPException(status_code=404, detail=f"Calculations not found for run {run_id}.")
    
    try:
        # Explicitly import openpyxl here to ensure it's available
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        with open(calc_file_path, "r", encoding="utf-8") as f:
            calculations = json.load(f)
        
        # Generate Excel file
        excel_path = run_dir / f"{run_id}_report.xlsx"
        
        # Convert Path to string for compatibility
        file_path_str = str(excel_path)
        
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            sheet = workbook["Sheet"]
            workbook.remove(sheet)
        
        # Create summary sheet
        summary_sheet = workbook.create_sheet("Summary")
        
        # Headers and styling
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=12, bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Report title
        summary_sheet['A1'] = f"GradeLens Report: {calculations.get('run_name', 'Unnamed Run')}"
        summary_sheet['A1'].font = title_font
        summary_sheet.merge_cells('A1:G1')
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Report date
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        summary_sheet.merge_cells('A2:G2')
        summary_sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Summary statistics
        summary_sheet['A4'] = "Overall Statistics"
        summary_sheet['A4'].font = header_font
        summary_sheet.merge_cells('A4:G4')
        
        row = 5
        summary_sheet[f'A{row}'] = "Total Students:"
        summary_sheet[f'B{row}'] = calculations.get("summary", {}).get("total_students", 0)
        
        row += 1
        summary_sheet[f'A{row}'] = "Overall GPA:"
        summary_sheet[f'B{row}'] = calculations.get("summary", {}).get("overall_gpa", 0)
        
        # Grade distribution headers
        row += 2
        summary_sheet[f'A{row}'] = "Grade Distribution"
        summary_sheet[f'A{row}'].font = header_font
        summary_sheet.merge_cells(f'A{row}:G{row}')
        
        row += 1
        headers = ["Grade", "Count", "Percentage"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            summary_sheet[f'{col}{row}'] = header
            summary_sheet[f'{col}{row}'].font = header_font
            summary_sheet[f'{col}{row}'].fill = header_fill
        
        # Grade distribution data
        grade_distribution = calculations.get("summary", {}).get("grade_distribution", {})
        total_grades = sum(grade_distribution.values())
        
        # Sort grades in logical order: A, B, C, D, F, W, Other
        grade_order = ["A", "B", "C", "D", "F", "W", "Other"]
        
        row += 1
        for grade in grade_order:
            count = grade_distribution.get(grade, 0)
            percentage = (count / total_grades * 100) if total_grades > 0 else 0
            
            summary_sheet[f'A{row}'] = grade
            summary_sheet[f'B{row}'] = count
            summary_sheet[f'C{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Course type analysis
        row += 2
        summary_sheet[f'A{row}'] = "Course Level Analysis"
        summary_sheet[f'A{row}'].font = header_font
        summary_sheet.merge_cells(f'A{row}:G{row}')
        
        row += 1
        headers = ["Course Level", "Students", "Average GPA", "Courses"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            summary_sheet[f'{col}{row}'] = header
            summary_sheet[f'{col}{row}'].font = header_font
            summary_sheet[f'{col}{row}'].fill = header_fill
        
        # Course type data
        class_types = calculations.get("class_types", {})
        for course_type, type_data in class_types.items():
            row += 1
            summary_sheet[f'A{row}'] = course_type
            summary_sheet[f'B{row}'] = type_data.get("total_students", 0)
            summary_sheet[f'C{row}'] = type_data.get("average_gpa", 0)
            summary_sheet[f'D{row}'] = ", ".join(type_data.get("courses", []))
        
        # Work-list students
        work_list_sheet = workbook.create_sheet("Work List Students")
        work_list_sheet['A1'] = "Students On Work List (GPA < 2.0)"
        work_list_sheet['A1'].font = title_font
        work_list_sheet.merge_cells('A1:F1')
        work_list_sheet['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        headers = ["Name", "Student ID", "GPA", "Courses", "Grades"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            work_list_sheet[f'{col}{row}'] = header
            work_list_sheet[f'{col}{row}'].font = header_font
            work_list_sheet[f'{col}{row}'].fill = header_fill
        
        work_list_students = calculations.get("improvement_lists", {}).get("work_list", [])
        for student in work_list_students:
            row += 1
            work_list_sheet[f'A{row}'] = student.get("name", "Unknown")
            work_list_sheet[f'B{row}'] = student.get("id", "")
            work_list_sheet[f'C{row}'] = student.get("gpa", 0)
            
            courses = [c.get("course", "") for c in student.get("courses", [])]
            work_list_sheet[f'D{row}'] = ", ".join(courses)
            
            grades = [f"{c.get('course')}: {c.get('grade')}" for c in student.get("courses", [])]
            work_list_sheet[f'E{row}'] = ", ".join(grades)
        
        # Good list students
        good_list = workbook.create_sheet("Good List Students")
        good_list['A1'] = "Good List Students (GPA ≥ 3.3)"
        good_list['A1'].font = title_font
        good_list.merge_cells('A1:F1')
        good_list['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        headers = ["Name", "Student ID", "GPA", "Courses", "Grades"]
        for i, header in enumerate(headers):
            col = chr(ord('A') + i)
            good_list[f'{col}{row}'] = header
            good_list[f'{col}{row}'].font = header_font
            good_list[f'{col}{row}'].fill = header_fill
        
        good_students = calculations.get("improvement_lists", {}).get("good_list", [])
        for student in good_students:
            row += 1
            good_list[f'A{row}'] = student.get("name", "Unknown")
            good_list[f'B{row}'] = student.get("id", "")
            good_list[f'C{row}'] = student.get("gpa", 0)
            
            courses = [c.get("course", "") for c in student.get("courses", [])]
            good_list[f'D{row}'] = ", ".join(courses)
            
            grades = [f"{c.get('course')}: {c.get('grade')}" for c in student.get("courses", [])]
            good_list[f'E{row}'] = ", ".join(grades)
        
        # Course sheets - FIXED THIS PART
        # Use course_list which is a list, or convert courses dictionary to a list
        course_list = calculations.get("course_list", [])
        if not course_list and "courses" in calculations:
            # If course_list is empty but courses exists as a dictionary, convert it
            if isinstance(calculations["courses"], dict):
                course_list = list(calculations["courses"].values())
        
        for i, course in enumerate(course_list):
            course_name = course.get("course_name", f"Course {i+1}")
            sheet_name = f"{course_name[:29]}"  # Excel sheet name length limit
            
            # Ensure sheet name is valid and unique
            sheet_name = re.sub(r'[\\/*\[\]:?]', '', sheet_name)  # Remove invalid chars
            if sheet_name in workbook.sheetnames:
                sheet_name = f"{sheet_name}_{i+1}"
                
            course_sheet = workbook.create_sheet(sheet_name)
            
            # Course header
            course_sheet['A1'] = f"Group: {course_name}"
            course_sheet['A1'].font = title_font
            course_sheet.merge_cells('A1:G1')
            course_sheet['A1'].alignment = Alignment(horizontal='center')
            
            row = 2
            course_sheet[f'A{row}'] = f"Type: {course.get('course_type', '')}"
            course_sheet[f'C{row}'] = f"Total Students: {course.get('total_students', 0)}"
            course_sheet[f'E{row}'] = f"Average GPA: {course.get('average_gpa', 0)}"
            
            # Grade distribution
            row += 2
            course_sheet[f'A{row}'] = "Grade Distribution"
            course_sheet[f'A{row}'].font = header_font
            
            row += 1
            headers = ["Grade", "Count"]
            for j, header in enumerate(headers):
                col = chr(ord('A') + j)
                course_sheet[f'{col}{row}'] = header
                course_sheet[f'{col}{row}'].font = header_font
                course_sheet[f'{col}{row}'].fill = header_fill
            
            grade_distribution = course.get("grade_distribution", {})
            for k, grade in enumerate(grade_order):
                if grade in grade_distribution:
                    count = grade_distribution.get(grade, 0)
                    row_idx = row + 1 + k
                    course_sheet[f'A{row_idx}'] = grade
                    course_sheet[f'B{row_idx}'] = count
            
            # Section information
            row += len(grade_order) + 2
            course_sheet[f'A{row}'] = "Sections"
            course_sheet[f'A{row}'].font = header_font
            
            row += 1
            headers = ["Section", "Students", "Average GPA"]
            for j, header in enumerate(headers):
                col = chr(ord('A') + j)
                course_sheet[f'{col}{row}'] = header
                course_sheet[f'{col}{row}'].font = header_font
                course_sheet[f'{col}{row}'].fill = header_fill
            
            for section in course.get("sections", []):
                row += 1
                course_sheet[f'A{row}'] = section.get("section_name", "")
                course_sheet[f'B{row}'] = section.get("student_count", 0)
                course_sheet[f'C{row}'] = section.get("average_gpa", 0)
        
        # FIX: Safe column width adjustment that properly handles merged cells
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Set default widths for columns
            for col_idx in range(1, 20):  # Support up to 20 columns
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 15
            
            # Then adjust based on content (safely)
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell, openpyxl.cell.Cell) and cell.value:  # Skip merged cells
                        col_letter = get_column_letter(cell.column)
                        try:
                            # Get the current width
                            current_width = worksheet.column_dimensions[col_letter].width
                            # Calculate new width based on content
                            content_length = len(str(cell.value)) + 4  # Add padding
                            # Use the larger of the current width or content length, max 50
                            worksheet.column_dimensions[col_letter].width = min(
                                max(current_width, content_length), 50
                            )
                        except:
                            pass  # Skip any problematic cells
        
        # Ensure the directory exists
        os.makedirs(os.path.dirname(file_path_str), exist_ok=True)
        
        # Save the workbook
        workbook.save(file_path_str)
        
        print(f"Excel file saved successfully to {file_path_str}")
        
        return FileResponse(
            path=file_path_str, 
            filename=f"GradeLens_Report_{calculations.get('run_name', 'Unnamed')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except ImportError as e:
        raise HTTPException(
            status_code=501,
            detail=f"Excel export is not available: {str(e)}"
        )
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel export error: {error_details}")
        raise HTTPException(status_code=500, detail=f"Error creating Excel report: {str(e)}")

# Z-score calculation for group courses
def z_scores_for_group_courses(courses_in_group, course_gpa_map):
    """
    Calculate Z-scores for courses within a group based on their GPAs.
    
    Args:
        courses_in_group: List of course codes (e.g., ['COMSC110', 'COMSC230', 'COMSC401'])
        course_gpa_map: Dictionary mapping course code to average GPA (e.g., {'COMSC110': 3.1, ...})
    
    Returns:
        Dictionary mapping course code to Z-score.
    """
    gpas = [course_gpa_map[code] for code in courses_in_group if code in course_gpa_map]
    if not gpas:
        return {}
    mean = sum(gpas) / len(gpas)
    std = (sum((g - mean) ** 2 for g in gpas) / len(gpas)) ** 0.5 if len(gpas) > 1 else 0.0
    z_scores = {}
    for code in courses_in_group:
        gpa = course_gpa_map.get(code)
        if gpa is not None:
            z_scores[code] = (gpa - mean) / std if std > 0 else 0.0
    return z_scores

# Main application entry point
app.include_router(router, prefix="/api")